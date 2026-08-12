"""
测试用例数据模型与导出器单元测试

覆盖：
- TestCase / TestStep / TestSuite 数据模型验证
- map_function_to_testcase 映射逻辑
- JsonExporter 导出
- ExcelExporter 导出
"""

import json
import os
import tempfile

import pytest

from models.test_case import (
    TestCase,
    TestStep,
    TestSuite,
    Traceability,
    map_function_to_testcase,
    map_analysis_to_testsuite,
)
from exporters.json_exporter import JsonExporter
from exporters.excel_exporter import ExcelExporter


# =============================================================================
# Fixtures
# =============================================================================


@pytest.fixture
def sample_step() -> TestStep:
    return TestStep(step=1, action="打开登录页面", expected="页面正常加载")


@pytest.fixture
def sample_trace() -> Traceability:
    return Traceability(
        requirement_source="电商平台PRD.md",
        requirement_section="3.1 用户登录",
        requirement_text="用户可通过已注册的用户名和密码登录系统",
    )


@pytest.fixture
def sample_case(sample_step, sample_trace) -> TestCase:
    return TestCase(
        case_id="TC-FUNC-001-01-01",
        title="正常登录-使用正确的用户名和密码登录",
        function_id="FUNC-001-01",
        function_name="正常登录流程",
        type="功能测试",
        priority="P0",
        precondition="已有注册账号 testuser/Test@123456",
        test_data={"username": "testuser", "password": "Test@123456"},
        steps=[sample_step],
        tags=["登录", "冒烟测试"],
        design_method="场景法",
        cleanup="测试完成后退出登录",
        traceability=sample_trace,
        status="待执行",
    )


@pytest.fixture
def sample_suite(sample_case) -> TestSuite:
    suite = TestSuite(
        suite_name="电商平台 - 用户认证 测试用例集",
        product_name="电商平台",
        module_name="用户认证",
    )
    suite.test_cases = [sample_case]
    suite.refresh_stats()
    return suite


# =============================================================================
# TestStep / TestCase 模型验证
# =============================================================================


class TestTestStep:
    def test_valid_step(self):
        step = TestStep(step=1, action="点击按钮", expected="弹窗出现")
        assert step.step == 1
        assert step.action == "点击按钮"

    def test_step_number_must_be_positive(self):
        with pytest.raises(Exception):
            TestStep(step=0, action="操作", expected="结果")

    def test_action_cannot_be_empty(self):
        with pytest.raises(Exception):
            TestStep(step=1, action="", expected="结果")

    def test_expected_cannot_be_empty(self):
        with pytest.raises(Exception):
            TestStep(step=1, action="操作", expected="")


class TestTestCaseValidation:
    def test_invalid_type_raises(self):
        with pytest.raises(ValueError, match="无效用例类型"):
            TestCase(
                case_id="TC-001",
                title="测试",
                function_id="FUNC-001",
                type="无效类型",
            )

    def test_invalid_priority_raises(self):
        with pytest.raises(ValueError, match="无效优先级"):
            TestCase(
                case_id="TC-001",
                title="测试",
                function_id="FUNC-001",
                priority="P9",
            )

    def test_invalid_status_raises(self):
        with pytest.raises(ValueError, match="无效状态"):
            TestCase(
                case_id="TC-001",
                title="测试",
                function_id="FUNC-001",
                status="未知状态",
            )

    def test_defaults(self):
        tc = TestCase(case_id="TC-001", title="测试", function_id="FUNC-001")
        assert tc.type == "功能测试"
        assert tc.priority == "P1"
        assert tc.status == "待执行"
        assert tc.precondition == "无"
        assert tc.test_data == {}
        assert tc.steps == []
        assert tc.tags == []
        assert tc.cleanup == ""
        assert tc.created_at != ""

    def test_to_flat_row(self, sample_case):
        row = sample_case.to_flat_row()
        assert row["用例ID"] == "TC-FUNC-001-01-01"
        assert row["用例标题"] == "正常登录-使用正确的用户名和密码登录"
        assert row["优先级"] == "P0"
        assert "步骤1" in row["测试步骤"]
        assert "testuser" in row["测试数据"]


class TestTestSuite:
    def test_refresh_stats(self, sample_case):
        suite = TestSuite(suite_name="测试套件")
        suite.test_cases = [sample_case]
        suite.refresh_stats()
        assert suite.total_cases == 1
        assert suite.p0_count == 1
        assert suite.p1_count == 0
        assert suite.p2_count == 0

    def test_multi_priority_stats(self):
        suite = TestSuite(suite_name="多优先级")
        suite.test_cases = [
            TestCase(case_id="TC-001", title="P0", function_id="F-1", priority="P0"),
            TestCase(case_id="TC-002", title="P0-2", function_id="F-2", priority="P0"),
            TestCase(case_id="TC-003", title="P1", function_id="F-3", priority="P1"),
            TestCase(case_id="TC-004", title="P2", function_id="F-4", priority="P2"),
            TestCase(case_id="TC-005", title="P1-2", function_id="F-5", priority="P1"),
        ]
        suite.refresh_stats()
        assert suite.total_cases == 5
        assert suite.p0_count == 2
        assert suite.p1_count == 2
        assert suite.p2_count == 1


# =============================================================================
# map_function_to_testcase 映射
# =============================================================================


class TestMapFunctionToTestcase:
    def test_basic_mapping(self):
        tc = map_function_to_testcase(
            func_id="FUNC-001-01",
            func_name="正常登录流程",
            func_desc="输入正确的用户名和密码，登录成功并跳转首页",
            func_priority="P0",
            acceptance_criteria=[
                "使用正确的用户名和密码登录，页面跳转到首页",
            ],
            test_suggestions=[{"method": "场景法", "suggestion": "Happy Path 用例"}],
            case_index=1,
            requirement_source="电商平台PRD.md",
            requirement_section="3.1",
            requirement_text="用户可通过已注册的用户名和密码登录系统",
        )

        assert tc.case_id == "TC-FUNC-001-01-01"
        assert tc.function_id == "FUNC-001-01"
        assert tc.function_name == "正常登录流程"
        assert tc.priority == "P0"
        assert tc.type == "功能测试"
        assert tc.design_method == "场景法"
        assert len(tc.steps) == 1
        assert tc.steps[0].action == "使用正确的用户名和密码登录，页面跳转到首页"
        assert tc.traceability.requirement_source == "电商平台PRD.md"
        assert tc.traceability.requirement_section == "3.1"

    def test_type_inference_security(self):
        tc = map_function_to_testcase(
            func_id="FUNC-002-01",
            func_name="权限验证",
            func_desc="验证用户权限和认证令牌安全性",
            func_priority="P0",
            acceptance_criteria=["验证通过"],
            test_suggestions=[],
        )
        assert tc.type == "安全测试"

    def test_type_inference_performance(self):
        tc = map_function_to_testcase(
            func_id="FUNC-003-01",
            func_name="并发下单",
            func_desc="高并发场景下的下单性能",
            func_priority="P1",
            acceptance_criteria=["QPS > 1000"],
            test_suggestions=[],
        )
        assert tc.type == "性能测试"

    def test_type_inference_api(self):
        tc = map_function_to_testcase(
            func_id="FUNC-004-01",
            func_name="API调用",
            func_desc="通过REST接口调用服务",
            func_priority="P1",
            acceptance_criteria=["返回200"],
            test_suggestions=[],
        )
        assert tc.type == "接口测试"

    def test_case_index_formatting(self):
        tc = map_function_to_testcase(
            func_id="FUNC-001-01",
            func_name="测试",
            func_desc="测试",
            func_priority="P1",
            acceptance_criteria=[],
            test_suggestions=[],
            case_index=5,
        )
        assert tc.case_id == "TC-FUNC-001-01-05"

    def test_p0_tags(self):
        tc = map_function_to_testcase(
            func_id="FUNC-001-01",
            func_name="支付",
            func_desc="核心支付流程",
            func_priority="P0",
            acceptance_criteria=["支付成功"],
            test_suggestions=[],
        )
        assert "冒烟测试" in tc.tags
        assert "核心流程" in tc.tags

    def test_non_p0_no_default_tags(self):
        tc = map_function_to_testcase(
            func_id="FUNC-001-01",
            func_name="帮助中心",
            func_desc="帮助",
            func_priority="P2",
            acceptance_criteria=["显示帮助"],
            test_suggestions=[],
        )
        assert tc.tags == []


# =============================================================================
# JsonExporter
# =============================================================================


class TestJsonExporter:
    def test_export_suite(self, sample_suite):
        exporter = JsonExporter(indent=2)
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "test_suite.json")
            result = exporter.export_suite(sample_suite, path)

            assert result == os.path.abspath(path)
            assert os.path.exists(path)

            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)

            assert data["suite_name"] == "电商平台 - 用户认证 测试用例集"
            assert data["total_cases"] == 1
            assert len(data["test_cases"]) == 1
            assert data["test_cases"][0]["case_id"] == "TC-FUNC-001-01-01"

    def test_export_cases_list(self, sample_case):
        exporter = JsonExporter()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "cases.json")
            result = exporter.export_cases([sample_case], path)

            assert os.path.exists(path)
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)

            assert isinstance(data, list)
            assert len(data) == 1
            assert data[0]["title"] == "正常登录-使用正确的用户名和密码登录"

    def test_compact_mode(self, sample_suite):
        exporter = JsonExporter(indent=None)
        json_str = exporter.export_to_string(sample_suite)
        assert "\n" not in json_str  # 紧凑模式无换行

    def test_auto_create_directory(self, sample_suite):
        exporter = JsonExporter()
        with tempfile.TemporaryDirectory() as tmpdir:
            nested = os.path.join(tmpdir, "a", "b", "c", "suite.json")
            result = exporter.export_suite(sample_suite, nested)
            assert os.path.exists(result)


# =============================================================================
# ExcelExporter
# =============================================================================


class TestExcelExporter:
    def test_export_creates_file(self, sample_suite):
        exporter = ExcelExporter()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "test_cases.xlsx")
            result = exporter.export_suite(sample_suite, path)

            assert result == os.path.abspath(path)
            assert os.path.exists(path)
            assert os.path.getsize(path) > 0

    def test_export_has_three_sheets(self, sample_suite):
        exporter = ExcelExporter()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "test_cases.xlsx")
            exporter.export_suite(sample_suite, path)

            from openpyxl import load_workbook
            wb = load_workbook(path)
            assert set(wb.sheetnames) == {"统计摘要", "用例列表", "步骤明细"}

    def test_summary_sheet_content(self, sample_suite):
        exporter = ExcelExporter()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "test_cases.xlsx")
            exporter.export_suite(sample_suite, path)

            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb["统计摘要"]

            assert ws.cell(row=1, column=1).value == "测试套件名称"
            assert ws.cell(row=1, column=2).value == sample_suite.suite_name
            assert ws.cell(row=5, column=2).value == 1  # total_cases

    def test_cases_sheet_header(self, sample_suite):
        exporter = ExcelExporter()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "test_cases.xlsx")
            exporter.export_suite(sample_suite, path)

            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb["用例列表"]

            assert ws.cell(row=1, column=1).value == "用例ID"
            assert ws.cell(row=2, column=1).value == "TC-FUNC-001-01-01"

    def test_steps_sheet_expands_steps(self, sample_suite):
        """步骤明细工作表应将每个步骤展开为一行"""
        exporter = ExcelExporter()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "test_cases.xlsx")
            exporter.export_suite(sample_suite, path)

            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb["步骤明细"]

            assert ws.cell(row=2, column=1).value == "TC-FUNC-001-01-01"
            assert ws.cell(row=2, column=3).value == 1  # 步骤序号
            assert "打开登录页面" in str(ws.cell(row=2, column=4).value)

    def test_freeze_panes(self, sample_suite):
        exporter = ExcelExporter()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "test_cases.xlsx")
            exporter.export_suite(sample_suite, path)

            from openpyxl import load_workbook
            wb = load_workbook(path)

            assert wb["用例列表"].freeze_panes == "A2"
            assert wb["步骤明细"].freeze_panes == "A2"

    def test_empty_suite(self):
        suite = TestSuite(suite_name="空套件")
        suite.refresh_stats()

        exporter = ExcelExporter()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "empty.xlsx")
            result = exporter.export_suite(suite, path)
            assert os.path.exists(result)

    def test_auto_create_directory(self, sample_suite):
        exporter = ExcelExporter()
        with tempfile.TemporaryDirectory() as tmpdir:
            nested = os.path.join(tmpdir, "deep", "nested", "cases.xlsx")
            result = exporter.export_suite(sample_suite, nested)
            assert os.path.exists(result)


# =============================================================================
# 端到端：从分析结果构建并导出
# =============================================================================


class TestEndToEnd:
    """模拟完整流程：分析结果 → TestSuite → 导出"""

    def test_analysis_to_export(self):
        # 模拟 RequirementAnalysisResult
        class MockOverview:
            product_name = "测试产品"
            module_name = "测试模块"
            total_functions = 1
            p0_count = 1
            p1_count = 0
            p2_count = 0

        class MockSubFunc:
            id = "FUNC-001-01"
            name = "正常流程"
            description = "核心正常流程"
            priority = "P0"
            acceptance_criteria = ["条件1", "条件2"]
            test_suggestions = [{"method": "场景法", "suggestion": "Happy Path"}]

        class MockFunc:
            id = "FUNC-001"
            name = "核心功能"
            description = "核心功能描述"
            priority = "P0"
            dependencies = []
            sub_functions = [MockSubFunc()]

            @staticmethod
            def model_dump():
                return {
                    "id": MockFunc.id,
                    "name": MockFunc.name,
                    "description": MockFunc.description,
                    "priority": MockFunc.priority,
                    "dependencies": MockFunc.dependencies,
                    "sub_functions": [
                        {
                            "id": MockSubFunc.id,
                            "name": MockSubFunc.name,
                            "description": MockSubFunc.description,
                            "priority": MockSubFunc.priority,
                            "acceptance_criteria": MockSubFunc.acceptance_criteria,
                            "test_suggestions": MockSubFunc.test_suggestions,
                        }
                    ],
                }

        class MockResult:
            overview = MockOverview()
            function_tree = [MockFunc()]

        result = MockResult()
        suite = map_analysis_to_testsuite(result, requirement_source="test.md")

        assert suite.suite_name == "测试产品 - 测试模块 测试用例集"
        assert suite.total_cases == 1
        assert suite.test_cases[0].case_id == "TC-FUNC-001-01-01"
        assert len(suite.test_cases[0].steps) == 2  # 2 个验收条件 → 2 个步骤
