"""
Excel 导出器 — 将 TestSuite / TestCase 导出为 .xlsx 文件

使用 openpyxl 库，支持：
- 自动列宽适配
- 表头冻结（首行）
- 优先级列条件着色（P0 红色 / P1 黄色 / P2 绿色）
- 统计摘要工作表
- 步骤明细工作表（拆分步骤，便于执行追踪）

特性：
- 自动创建输出目录
- 即使测试数据为空也能正常导出
"""

import os
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models.test_case import TestCase, TestSuite


# =============================================================================
# 样式常量
# =============================================================================

HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_FONT = Font(name="微软雅黑", size=10)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

P0_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 浅红
P1_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 浅黄
P2_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 浅绿

SUMMARY_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUMMARY_LABEL_FONT = Font(name="微软雅黑", bold=True, size=11)
SUMMARY_VALUE_FONT = Font(name="微软雅黑", size=11)

# 用例表列定义（映射 TestCase.to_flat_row() 键）
CASE_COLUMNS = [
    ("用例ID", 15),
    ("用例标题", 35),
    ("关联功能ID", 18),
    ("功能名称", 20),
    ("用例类型", 12),
    ("优先级", 8),
    ("前置条件", 30),
    ("测试数据", 25),
    ("测试步骤", 50),
    ("标签", 20),
    ("设计方法", 16),
    ("清理步骤", 25),
    ("需求来源", 20),
    ("需求章节", 20),
    ("状态", 10),
    ("创建时间", 20),
]

# 步骤明细列定义
STEP_COLUMNS = [
    ("用例ID", 18),
    ("用例标题", 35),
    ("步骤序号", 10),
    ("操作", 50),
    ("预期结果", 50),
    ("优先级", 8),
    ("状态", 10),
]


class ExcelExporter:
    """
    Excel 格式导出器

    Usage:
        exporter = ExcelExporter()
        path = exporter.export_suite(suite, "/output/test_cases.xlsx")
    """

    def __init__(self) -> None:
        pass

    @staticmethod
    def _ensure_dir(file_path: str) -> None:
        parent = os.path.dirname(file_path)
        if parent:
            os.makedirs(parent, exist_ok=True)

    @staticmethod
    def _apply_header_style(ws, row: int, col_count: int) -> None:
        """给表头行应用统一样式"""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

    @staticmethod
    def _apply_cell_style(ws, row: int, col_count: int, priority_col: int = -1) -> None:
        """给数据行应用统一样式，并根据优先级着色"""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

        # 优先级着色
        if priority_col > 0:
            priority_cell = ws.cell(row=row, column=priority_col)
            p_val = str(priority_cell.value or "")
            if p_val == "P0":
                priority_cell.fill = P0_FILL
            elif p_val == "P1":
                priority_cell.fill = P1_FILL
            elif p_val == "P2":
                priority_cell.fill = P2_FILL

    @staticmethod
    def _auto_column_width(ws, col_defs) -> None:
        """根据列定义自动设置列宽"""
        for col_idx, (_, width) in enumerate(col_defs, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # =========================================================================
    # 工作表 1：统计摘要
    # =========================================================================

    def _create_summary_sheet(self, wb: Workbook, suite: TestSuite) -> None:
        """创建统计摘要工作表"""
        ws = wb.active
        ws.title = "统计摘要"

        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        summary_data = [
            ("测试套件名称", suite.suite_name),
            ("产品名称", suite.product_name),
            ("模块名称", suite.module_name),
            ("", ""),
            ("用例总数", suite.total_cases),
            ("P0 用例数", suite.p0_count),
            ("P1 用例数", suite.p1_count),
            ("P2 用例数", suite.p2_count),
            ("", ""),
            ("生成时间", suite.generated_at),
            ("导出时间", now_str),
        ]

        # 写入数据
        for i, (label, value) in enumerate(summary_data, start=1):
            label_cell = ws.cell(row=i, column=1, value=label)
            value_cell = ws.cell(row=i, column=2, value=value)

            if label:
                label_cell.font = SUMMARY_LABEL_FONT
            value_cell.font = SUMMARY_VALUE_FONT

        # 列宽
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 40

        # P0/P1/P2 统计行着色
        for row in (6, 7, 8):
            p_cell = ws.cell(row=row, column=1)
            v_cell = ws.cell(row=row, column=2)
            val = str(v_cell.value or "")
            fill = None
            if "P0" in val:
                fill = P0_FILL
            elif "P1" in val:
                fill = P1_FILL
            elif "P2" in val:
                fill = P2_FILL
            if fill:
                p_cell.fill = fill
                v_cell.fill = fill

    # =========================================================================
    # 工作表 2：用例列表
    # =========================================================================

    def _create_cases_sheet(self, wb: Workbook, suite: TestSuite) -> None:
        """创建用例列表工作表"""
        ws = wb.create_sheet(title="用例列表")

        # 表头
        for col_idx, (col_name, _) in enumerate(CASE_COLUMNS, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        self._apply_header_style(ws, 1, len(CASE_COLUMNS))

        # 优先级列索引（CASE_COLUMNS 中"优先级"在第 6 列）
        priority_col = next(
            (i + 1 for i, (c, _) in enumerate(CASE_COLUMNS) if c == "优先级"), -1
        )

        # 数据行
        for row_idx, tc in enumerate(suite.test_cases, start=2):
            flat = tc.to_flat_row()
            for col_idx, (col_name, _) in enumerate(CASE_COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=flat.get(col_name, ""))
            self._apply_cell_style(ws, row_idx, len(CASE_COLUMNS), priority_col)

        # 列宽 + 冻结表头
        self._auto_column_width(ws, CASE_COLUMNS)
        ws.freeze_panes = "A2"

    # =========================================================================
    # 工作表 3：步骤明细
    # =========================================================================

    def _create_steps_sheet(self, wb: Workbook, suite: TestSuite) -> None:
        """创建步骤明细工作表（每个步骤一行，便于逐步骤执行追踪）"""
        ws = wb.create_sheet(title="步骤明细")

        # 表头
        for col_idx, (col_name, _) in enumerate(STEP_COLUMNS, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        self._apply_header_style(ws, 1, len(STEP_COLUMNS))

        priority_col = next(
            (i + 1 for i, (c, _) in enumerate(STEP_COLUMNS) if c == "优先级"), -1
        )

        row_idx = 2
        for tc in suite.test_cases:
            if not tc.steps:
                # 无步骤的用例也占一行
                for col_idx, (col_name, _) in enumerate(STEP_COLUMNS, start=1):
                    mapping = {
                        "用例ID": tc.case_id,
                        "用例标题": tc.title,
                        "步骤序号": "-",
                        "操作": "（无步骤）",
                        "预期结果": "-",
                        "优先级": tc.priority,
                        "状态": tc.status,
                    }
                    ws.cell(row=row_idx, column=col_idx, value=mapping.get(col_name, ""))
                self._apply_cell_style(ws, row_idx, len(STEP_COLUMNS), priority_col)
                row_idx += 1
                continue

            for step in tc.steps:
                for col_idx, (col_name, _) in enumerate(STEP_COLUMNS, start=1):
                    mapping = {
                        "用例ID": tc.case_id,
                        "用例标题": tc.title,
                        "步骤序号": step.step,
                        "操作": step.action,
                        "预期结果": step.expected,
                        "优先级": tc.priority,
                        "状态": tc.status,
                    }
                    ws.cell(row=row_idx, column=col_idx, value=mapping.get(col_name, ""))
                self._apply_cell_style(ws, row_idx, len(STEP_COLUMNS), priority_col)
                row_idx += 1

        self._auto_column_width(ws, STEP_COLUMNS)
        ws.freeze_panes = "A2"

    # =========================================================================
    # 公共 API
    # =========================================================================

    def export_suite(self, suite: TestSuite, output_path: str) -> str:
        """
        导出 TestSuite 为 Excel 文件（含三个工作表）

        Args:
            suite: 测试套件实例
            output_path: 输出 .xlsx 文件路径

        Returns:
            输出文件的绝对路径
        """
        self._ensure_dir(output_path)

        wb = Workbook()

        # 三个工作表
        self._create_summary_sheet(wb, suite)
        self._create_cases_sheet(wb, suite)
        self._create_steps_sheet(wb, suite)

        wb.save(output_path)
        return os.path.abspath(output_path)
